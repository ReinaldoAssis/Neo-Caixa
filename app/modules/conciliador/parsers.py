from __future__ import annotations

import csv
import tempfile
from pathlib import Path

from app.modules.conciliador.constants import (
    empty_categories_posto,
    empty_categories_restaurante,
    CATEGORIAS_RESTAURANTE,
    parse_money,
    normalize_text,
)
from app.modules.conciliador.config_posto import (
    default_config,
    config_category_keys,
    build_system_map,
    build_premmia_map,
    build_pagbank_map,
    INFO_FIELDS_DEFAULT,
)

OLE2_MAGIC = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"


# ─── Posto: Caixa CSV ────────────────────────────────────────────

INFO_FIELDS = INFO_FIELDS_DEFAULT


def parse_caixa_csv(file_content: bytes, filename: str, config: dict | None = None) -> dict:
    temp_path = Path(tempfile.gettempdir()) / f"caixa_{filename}"
    temp_path.write_bytes(file_content)
    try:
        return _parse_caixa_csv(temp_path, config or default_config())
    finally:
        try:
            temp_path.unlink()
        except OSError:
            pass


def _parse_caixa_csv(path: Path, config: dict) -> dict:
    system_map = build_system_map(config)
    keys = config_category_keys(config)

    with open(path, "r", encoding="latin-1", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=";"))

    first_lines = [";".join(row) for row in rows[:5]]
    if not any("CAIXA GERAL" in line.upper() for line in first_lines):
        raise ValueError("O arquivo selecionado nao parece ser um CSV do CAIXA GERAL.")

    header_index = None
    for idx, row in enumerate(rows):
        normalized = [normalize_text(cell) for cell in row]
        if len(normalized) >= 2 and normalized[0] == "ENTRADAS" and normalized[1].startswith("SA"):
            header_index = idx
            break
    if header_index is None:
        raise ValueError("Nao encontrei a secao FINANCEIRO com o cabecalho Entradas/Saidas.")

    categorias = empty_categories_posto(keys)
    info = {"sangria": 0.0, "notas_a_prazo": 0.0, "despesas": 0.0}
    detected: list[dict] = []
    not_found: list[dict] = []

    for row in rows[header_index + 1:]:
        if not row:
            continue
        if normalize_text(row[0]).startswith("SUBTOTAL"):
            break
        if len(row) < 4:
            continue
        name = normalize_text(row[2])
        value_text = row[3].strip() if len(row) > 3 else ""
        if not name or not value_text:
            continue
        value = parse_money(value_text)
        if name in system_map:
            key = system_map[name]
            categorias[key]["sistema"] = round(categorias[key]["sistema"] + value, 2)
            detected.append({"nome": name, "categoria": key, "valor": value})
        elif name in INFO_FIELDS:
            info[INFO_FIELDS[name]] = value
            detected.append({"nome": name, "categoria": INFO_FIELDS[name], "valor": value})
        else:
            not_found.append({"nome": name, "valor": value})

    return {
        "categorias": categorias,
        "sangria": info["sangria"],
        "notas_a_prazo": info["notas_a_prazo"],
        "despesas": info["despesas"],
        "detected": detected,
        "nao_reconhecidos": not_found,
        "total_saidas": round(sum(item["valor"] for item in detected), 2),
    }


# ─── Posto: PagBank CSV ──────────────────────────────────────────

def _pagbank_normalized_key(bandeira: str, forma: str) -> tuple[str, str]:
    brand = normalize_text(bandeira)
    method = normalize_text(forma)
    if method == "PIX" or (not brand and method == "PIX"):
        return ("", "PIX")
    if method.startswith("DEBIT"):
        method = "DEBITO"
    elif method.startswith("CREDIT"):
        method = "CREDITO"
    if brand in {"AMERICAN EXPRESS", "AMEX"}:
        brand = "AMEX"
    return (brand, method)


def parse_pagbank_csv(file_content: bytes, filename: str, config: dict | None = None) -> dict:
    temp_path = Path(tempfile.gettempdir()) / f"pagbank_{filename}"
    temp_path.write_bytes(file_content)
    try:
        return _parse_pagbank_csv(temp_path, config or default_config())
    finally:
        try:
            temp_path.unlink()
        except OSError:
            pass


def _parse_pagbank_csv(path: Path, config: dict) -> dict:
    pagbank_map = build_pagbank_map(config)
    keys = config_category_keys(config)

    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        if not reader.fieldnames or "Código da Transação" not in reader.fieldnames:
            raise ValueError("O arquivo selecionado nao parece ser o CSV PagBank esperado.")

        categorias = empty_categories_posto(keys)
        approved = 0
        ignored = 0
        nao_reconhecidos: dict[str, int] = {}
        for row in reader:
            if normalize_text(row.get("Status")) != "APROVADA":
                ignored += 1
                continue
            raw = _pagbank_normalized_key(row.get("Bandeira", ""), row.get("Forma de Pagamento", ""))
            key = pagbank_map.get(raw)
            if not key or key not in categorias:
                ignored += 1
                combo = f"{raw[0]}/{raw[1]}".strip("/")
                nao_reconhecidos[combo] = nao_reconhecidos.get(combo, 0) + 1
                continue
            value = parse_money(row.get("Valor Bruto", "0"))
            categorias[key]["site"] = round(categorias[key]["site"] + value, 2)
            approved += 1

    return {
        "categorias": categorias,
        "registros_aprovados": approved,
        "registros_ignorados": ignored,
        "nao_reconhecidos": [
            {"combo": k, "quantidade": v} for k, v in sorted(nao_reconhecidos.items())
        ],
    }


# ─── Posto: Premmia XLS/XLSX ─────────────────────────────────────

ZIP_MAGIC = b"PK\x03\x04"


def detectar_formato_premmia(content: bytes) -> str:
    if content[:8] == OLE2_MAGIC:
        return "xls"
    if content[:4] == ZIP_MAGIC:
        return "xlsx"
    return "csv"


def _rows_from_xls(content: bytes) -> list[dict]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("Instale xlrd para processar arquivos Premmia.") from exc

    temp = tempfile.NamedTemporaryFile(suffix=".xls", delete=False)
    try:
        temp.write(content)
        temp.close()
        book = xlrd.open_workbook(temp.name)
    finally:
        try:
            Path(temp.name).unlink()
        except OSError:
            pass

    if "Conferência" in book.sheet_names():
        sheet = book.sheet_by_name("Conferência")
    elif "Conferencia" in book.sheet_names():
        sheet = book.sheet_by_name("Conferencia")
    else:
        sheet = book.sheet_by_index(0)

    grid = [
        [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        for r in range(sheet.nrows)
    ]
    return _records_from_grid(grid)


def _rows_from_xlsx(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Instale openpyxl para processar arquivos Premmia .xlsx.") from exc

    temp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        temp.write(content)
        temp.close()
        book = openpyxl.load_workbook(temp.name, data_only=True, read_only=True)
    finally:
        try:
            Path(temp.name).unlink()
        except OSError:
            pass

    sheet = None
    for name in ("Conferência", "Conferencia"):
        if name in book.sheetnames:
            sheet = book[name]
            break
    if sheet is None:
        sheet = book[book.sheetnames[0]]

    grid = [
        ["" if v is None else v for v in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    return _records_from_grid(grid)


def _records_from_grid(grid: list[list]) -> list[dict]:
    header_row = None
    headers: list[str] = []
    for idx, row in enumerate(grid):
        values = [str(v).strip() for v in row]
        normalized = [normalize_text(v) for v in values]
        if "CPF" in normalized and "STATUS" in normalized:
            header_row = idx
            headers = values
            break
    if header_row is None:
        raise ValueError("Nao encontrei o cabecalho esperado no arquivo Premmia.")

    records = []
    for row in grid[header_row + 1:]:
        if not any(str(v).strip() for v in row):
            continue
        record = {}
        for col_idx, header in enumerate(headers):
            record[header] = row[col_idx] if col_idx < len(row) else ""
        records.append(record)
    return records


def _premmia_get(record: dict, wanted: str) -> object:
    wanted_norm = normalize_text(wanted)
    for key, value in record.items():
        if normalize_text(key) == wanted_norm:
            return value
    return ""


def parse_premmia_file(file_content: bytes, filename: str, config: dict | None = None) -> dict:
    config = config or default_config()
    premmia_map = build_premmia_map(config)
    keys = config_category_keys(config)

    formato = detectar_formato_premmia(file_content)
    if formato == "xls":
        records = _rows_from_xls(file_content)
    elif formato == "xlsx":
        records = _rows_from_xlsx(file_content)
    else:
        raise ValueError("O relatorio Premmia deve ser um arquivo Excel .xls ou .xlsx valido.")

    categorias = empty_categories_posto(keys)
    processed = 0
    ignored = 0
    nao_reconhecidos: dict[str, int] = {}
    lancamentos: list[dict] = []

    for record in records:
        status = normalize_text(_premmia_get(record, "Status"))
        forma_raw = str(_premmia_get(record, "Forma de Pagamento")).strip()
        forma = normalize_text(forma_raw)
        valor = parse_money(_premmia_get(record, "Valor líquido"))
        nome = str(_premmia_get(record, "Nome")).strip()
        cpf = str(_premmia_get(record, "CPF")).strip()
        data_hora = str(
            _premmia_get(record, "Data/Hora da transação")
            or _premmia_get(record, "Data/Hora da transacao")
        ).strip()

        key = premmia_map.get(forma)
        aceito = bool(key and key in categorias and status == "PROCESSADA")

        lancamentos.append({
            "cpf": cpf,
            "nome": nome,
            "forma": forma_raw,
            "valor": valor,
            "data_hora": data_hora,
            "status": str(_premmia_get(record, "Status")).strip(),
            "categoria": key if aceito else None,
            "aceito": aceito,
        })

        if status != "PROCESSADA":
            ignored += 1
            continue
        if not aceito:
            ignored += 1
            if forma:
                nao_reconhecidos[forma] = nao_reconhecidos.get(forma, 0) + 1
            continue
        categorias[key]["site"] = round(categorias[key]["site"] + valor, 2)
        processed += 1

    return {
        "formato": formato,
        "categorias": categorias,
        "transacoes_processadas": processed,
        "transacoes_ignoradas": ignored,
        "lancamentos": lancamentos,
        "nao_reconhecidos": [
            {"forma": k, "quantidade": v} for k, v in sorted(nao_reconhecidos.items())
        ],
    }


# ─── Restaurante: PagBank CSV ────────────────────────────────────

RESTAURANTE_CATEGORY_MAP: dict[tuple[str, str], str] = {
    ("PIX", "PIX"): "PIX",
    ("ELO", "DEBITO"): "ELO_DEBITO",
    ("ELO", "CREDITO"): "ELO_CR",
    ("MASTERCARD", "DEBITO"): "MAESTRO",
    ("MASTERCARD", "CREDITO"): "MASTERCARD",
    ("VISA", "DEBITO"): "VC_ELECTRON",
    ("VISA", "CREDITO"): "VISA",
    ("AMEX", "CREDITO"): "AMEX",
    ("AMEX", "AMBOS"): "AMEX",
}


def _restaurante_category_key(bandeira: str, forma: str) -> tuple[str, str]:
    brand = normalize_text(bandeira)
    method = normalize_text(forma)
    if not brand and method == "PIX":
        return ("PIX", "PIX")
    if brand == "ELO":
        return ("ELO", "DEBITO" if method.startswith("DEBIT") else "CREDITO")
    if brand == "MASTERCARD":
        return ("MASTERCARD", "DEBITO" if method.startswith("DEBIT") else "CREDITO")
    if brand == "VISA":
        return ("VISA", "DEBITO" if method.startswith("DEBIT") else "CREDITO")
    if brand == "AMEX":
        return ("AMEX", "CREDITO" if method.startswith("CREDIT") else method)
    return (brand, method)


def parse_restaurante_pagbank_csv(
    file_content: bytes,
    filename: str,
    hora_ini: str | None = None,
    hora_fim: str | None = None,
) -> dict:
    temp_path = Path(tempfile.gettempdir()) / f"rest_pagbank_{filename}"
    temp_path.write_bytes(file_content)
    try:
        return _parse_restaurante_pagbank_csv(temp_path, hora_ini, hora_fim)
    finally:
        try:
            temp_path.unlink()
        except OSError:
            pass


def parse_restaurante_pagbank_split(
    file_content: bytes,
    filename: str,
    split_time: str,
) -> dict:
    """Divide um unico relatorio PagBank em dois turnos por horario.

    Transacoes com horario < split_time somam no turno 1, o restante no turno 2.
    """
    from datetime import datetime

    try:
        split = datetime.strptime(split_time.strip(), "%H:%M").time()
    except ValueError:
        raise ValueError("Horario de divisao invalido. Use o formato HH:MM.")

    temp_path = Path(tempfile.gettempdir()) / f"rest_split_{filename}"
    temp_path.write_bytes(file_content)
    try:
        cat1 = empty_categories_restaurante()
        cat2 = empty_categories_restaurante()
        approved1 = 0
        approved2 = 0
        with open(temp_path, "r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=";")
            if not reader.fieldnames or "Código da Transação" not in reader.fieldnames:
                raise ValueError("O arquivo selecionado nao parece ser o CSV PagBank esperado.")
            for row in reader:
                if normalize_text(row.get("Status", "")) != "APROVADA":
                    continue
                raw_key = _restaurante_category_key(
                    row.get("Bandeira", ""), row.get("Forma de Pagamento", "")
                )
                key = RESTAURANTE_CATEGORY_MAP.get(raw_key)
                if not key or key == "DINHEIRO":
                    continue
                value = parse_money(row.get("Valor Bruto", "0"))
                target = cat1
                is_turno1 = True
                dt_str = row.get("Data da Transação", "")
                if dt_str:
                    try:
                        tx_time = datetime.strptime(dt_str.strip(), "%d/%m/%Y %H:%M").time()
                        is_turno1 = tx_time < split
                    except ValueError:
                        is_turno1 = True
                if is_turno1:
                    cat1[key]["real"] = round(cat1[key]["real"] + value, 2)
                    approved1 += 1
                else:
                    cat2[key]["real"] = round(cat2[key]["real"] + value, 2)
                    approved2 += 1
        return {
            "turno1": {"categorias": cat1, "registros_aprovados": approved1},
            "turno2": {"categorias": cat2, "registros_aprovados": approved2},
            "split_time": split_time,
        }
    finally:
        try:
            temp_path.unlink()
        except OSError:
            pass


def _parse_restaurante_pagbank_csv(
    path: Path,
    hora_ini: str | None = None,
    hora_fim: str | None = None,
) -> dict:
    from datetime import datetime

    t_ini = None
    t_fim = None
    if hora_ini and hora_fim:
        try:
            t_ini = datetime.strptime(hora_ini.strip(), "%H:%M").time()
            t_fim = datetime.strptime(hora_fim.strip(), "%H:%M").time()
        except ValueError:
            raise ValueError("Horario invalido. Use o formato HH:MM.")

    categorias = empty_categories_restaurante()
    approved = 0

    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        if not reader.fieldnames or "Código da Transação" not in reader.fieldnames:
            raise ValueError("O arquivo selecionado nao parece ser o CSV PagBank esperado.")

        for row in reader:
            if normalize_text(row.get("Status", "")) != "APROVADA":
                continue
            raw_key = _restaurante_category_key(
                row.get("Bandeira", ""), row.get("Forma de Pagamento", "")
            )
            key = RESTAURANTE_CATEGORY_MAP.get(raw_key)
            if not key or key == "DINHEIRO":
                continue

            if t_ini and t_fim:
                dt_str = row.get("Data da Transação", "")
                if dt_str:
                    try:
                        tx_dt = datetime.strptime(dt_str.strip(), "%d/%m/%Y %H:%M")
                        tx_time = tx_dt.time()
                        if not (t_ini <= tx_time <= t_fim):
                            continue
                    except ValueError:
                        pass

            value = parse_money(row.get("Valor Bruto", "0"))
            categorias[key]["real"] = round(categorias[key]["real"] + value, 2)
            approved += 1

    return {
        "categorias": categorias,
        "registros_aprovados": approved,
    }
